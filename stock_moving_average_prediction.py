import numpy as np
import pandas as pd
import yfinance as yf
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import classification_report
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Configure logging
logging.basicConfig(level=logging.INFO)

# Function to calculate RSI


def calculate_rsi(data, window=14):
    delta = data['Close'].diff(1)
    gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

# Fetch data


def fetch_data(ticker):
    logging.info("Fetching data...")
    data = yf.download(ticker, start="2020-01-01", end="2023-01-01")
    logging.info("Data fetched successfully.")
    return data


# Main execution
if __name__ == "__main__":
    # Fetch stock data
    ticker = 'AAPL'
    data = fetch_data(ticker)

    # Feature Engineering
    data['MA_50'] = data['Close'].rolling(window=50).mean()
    data['MA_200'] = data['Close'].rolling(window=200).mean()
    data['RSI'] = calculate_rsi(data)
    data.dropna(inplace=True)

    # Create target variable
    data['Target'] = np.where(data['Close'].shift(-1) > data['Close'], 1, 0)

    # Prepare features and target variable
    features = ['Close', 'MA_50', 'MA_200', 'RSI']
    X = data[features]
    y = data['Target']

    # Split data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42)

    # Train model using Grid Search for hyperparameter tuning
    model = RandomForestClassifier()
    param_grid = {'n_estimators': [50, 100], 'max_depth': [None, 10, 20]}
    grid_search = GridSearchCV(model, param_grid, cv=5)
    grid_search.fit(X_train, y_train)

    logging.info("Model trained with best parameters.")
    best_model = grid_search.best_estimator_

    # Make predictions
    predictions = best_model.predict(X_test)

    # Evaluate model
    report = classification_report(y_test, predictions)
    logging.info(f"Classification Report:\n{report}")

    # Add predictions to the DataFrame
    data.loc[data.index.isin(X_test.index), 'Predicted'] = predictions

    # Create descriptions for predictions
    data['Description'] = np.where(data['Predicted'] == 1,
                                   'The model predicts a price increase tomorrow, indicating positive market sentiment.',
                                   'The model predicts no price increase or a decrease tomorrow, indicating negative or neutral market sentiment.')

    # Select last 20 rows
    last_20_rows = data.tail(20)

    # Create an Excel writer object
    excel_filename = f"stock_predictions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        last_20_rows.to_excel(writer, sheet_name='Predictions', index=True)

        # Access the workbook and sheet
        workbook = writer.book
        sheet = writer.sheets['Predictions']

        # Define fill colors for green and yellow
        green_fill = PatternFill(start_color='00FF00',
                                 end_color='00FF00', fill_type='solid')
        yellow_fill = PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Apply conditional formatting
        # +2 to account for header row
        for row in range(2, last_20_rows.shape[0] + 2):
            # Get the predicted value
            predicted_value = last_20_rows['Predicted'].iloc[row - 2]

            if predicted_value == 1:
                # Apply green fill for buy predictions
                for col in range(1, last_20_rows.shape[1] + 1):
                    sheet.cell(row=row, column=col).fill = green_fill
            elif predicted_value == 0:
                # Apply yellow fill for hold/sell predictions
                for col in range(1, last_20_rows.shape[1] + 1):
                    sheet.cell(row=row, column=col).fill = yellow_fill

    logging.info(
        f"Excel file '{excel_filename}' generated successfully with predictions.")
