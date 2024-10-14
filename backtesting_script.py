import numpy as np
import pandas as pd
import yfinance as yf
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import classification_report
import logging
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Configure logging
logging.basicConfig(level=logging.INFO)

# Function to calculate RSI


def calculate_rsi(data, window=14):
    delta = data['Close'].diff(1)
    gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

# Fetch data


def fetch_data(ticker):
    logging.info("Fetching data...")
    data = yf.download(ticker, start="2020-01-01", end="2023-01-01")
    logging.info("Data fetched successfully.")
    return data

# Backtest strategy


def backtest_strategy(data, initial_balance=10000):
    balance = initial_balance
    shares = 0
    trade_summary = []

    for index, row in data.iterrows():
        trade_time = row.name  # Capture date and time of trade

        # Risk Management: Ensure enough balance to buy shares
        if row['Predicted'] == 1 and balance >= row['Close']:
            # Allow fractional shares if necessary
            shares = balance / row['Close']
            balance -= shares * row['Close']
            trade_summary.append({
                'Trade': 'Buy',
                'Date': trade_time.date(),
                'Time': trade_time.time(),
                'Price': row['Close'],
                'Shares': shares,
                'Balance': balance,
                'Profit/Loss': 0  # Placeholder for future profit/loss calculation
            })

        elif row['Predicted'] == 0 and shares > 0:
            # Ensure the price is treated as float
            price = float(row['Close'])
            trade_summary.append({
                'Trade': 'Sell',
                'Date': trade_time.date(),
                'Time': trade_time.time(),
                'Price': price,
                'Shares': shares,
                'Balance': balance + shares * price,
                'Profit/Loss': (price - float(trade_summary[-1]['Price'])) * shares
            })
            balance += shares * price  # Update balance after selling
            shares = 0

    if not trade_summary:
        logging.warning("No trades were executed during backtesting.")
        return []

    final_balance = balance + \
        (shares * data.iloc[-1]['Close'] if shares > 0 else 0)
    profit_loss = final_balance - initial_balance
    logging.info(f"Final portfolio value: ${final_balance:.2f}")
    logging.info(f"Profit/Loss: ${profit_loss:.2f}")

    return trade_summary


# Main execution
if __name__ == "__main__":
    try:
        # Fetch stock data
        ticker = 'AAPL'
        data = fetch_data(ticker)

        # Feature Engineering
        data['MA_50'] = data['Close'].rolling(window=50).mean()
        data['MA_200'] = data['Close'].rolling(window=200).mean()
        data['RSI'] = calculate_rsi(data)
        data.dropna(inplace=True)

        # Create target variable
        data['Target'] = np.where(
            data['Close'].shift(-1) > data['Close'], 1, 0)

        # Prepare features and target variable
        features = ['Close', 'MA_50', 'MA_200', 'RSI']
        X = data[features]
        y = data['Target']

        # Split data into training and test sets
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42)

        # Train model using Grid Search for hyperparameter tuning
        model = RandomForestClassifier()
        param_grid = {'n_estimators': [50, 100], 'max_depth': [None, 10, 20]}
        grid_search = GridSearchCV(model, param_grid, cv=5)
        grid_search.fit(X_train, y_train)

        logging.info("Model trained with best parameters.")
        best_model = grid_search.best_estimator_

        # Make predictions
        predictions = best_model.predict(X_test)

        # Evaluate model
        report = classification_report(y_test, predictions)
        logging.info(f"Classification Report:\n{report}")

        # Check distribution of buy (1) and sell (0) signals
        logging.info(f"Buy signals: {sum(predictions == 1)}")
        logging.info(f"Sell signals: {sum(predictions == 0)}")

        # Add predictions to the DataFrame
        data.loc[data.index.isin(X_test.index), 'Predicted'] = predictions

        # Backtesting only on the rows with predictions
        backtest_data = data.dropna(subset=['Predicted'])
        trade_summary = backtest_strategy(backtest_data)

        # Create a DataFrame for trade summary
        trade_summary_df = pd.DataFrame(trade_summary)

        # Save results to Excel
        excel_filename = f"excels/stock_backtesting_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        if not os.path.exists('excels'):
            os.makedirs('excels')

        # Create Excel workbook and sheets
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Trade Summary"

        # Write headers
        headers = trade_summary_df.columns.tolist()
        summary_sheet.append(headers)

        # Apply color scheme and write data
        for index, row in trade_summary_df.iterrows():
            summary_sheet.append(row.values.tolist())
            fill = PatternFill(start_color="00FF00" if row['Trade'] == 'Buy' else "FF0000",
                               end_color="00FF00" if row['Trade'] == 'Buy' else "FF0000", fill_type="solid")
            for cell in summary_sheet[index + 2]:  # +2 for header row
                cell.fill = fill

        # Save workbook
        wb.save(excel_filename)
        logging.info(f"Trade summary saved to {excel_filename}")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
